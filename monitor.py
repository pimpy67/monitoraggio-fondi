"""
monitor.py - Script principale di monitoraggio fondi
=====================================================
Orchestrazione del sistema:
1. Legge file Excel con lista fondi
2. Recupera NAV per ogni fondo
3. Calcola indicatori tecnici
4. Genera segnali
5. Invia alert
6. Aggiorna Excel e dashboard
"""

import os
import sys
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
import time
from decimal import Decimal
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError

# Carica variabili .env se presenti (utile quando monitor.py viene eseguito direttamente)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Import moduli locali
from data_fetcher import FundDataFetcher
from technical_analysis import TechnicalAnalyzer
from alerts import AlertSystem
from database import PriceDatabase


# Log globale degli errori consultabile via /api/monitor-log
monitor_log = []


def add_log(message: str):
    """Aggiunge un messaggio al log globale"""
    entry = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    monitor_log.append(entry)
    # Tieni solo gli ultimi 200 messaggi
    if len(monitor_log) > 200:
        monitor_log.pop(0)
    print(entry)


class FundMonitor:
    """Sistema principale di monitoraggio fondi"""

    def __init__(self, excel_path: str = 'fondi_monitoraggio.xlsx'):
        """
        Inizializza il monitor
        
        Args:
            excel_path: Percorso al file Excel master
        """
        self.excel_path = excel_path
        self.data_fetcher = FundDataFetcher()
        self.analyzer = TechnicalAnalyzer()
        self.alert_system = AlertSystem()

        # Database PostgreSQL per storico prezzi (persistente su Railway)
        self.db = PriceDatabase()

        # Fallback: storage locale per dati storici (se DB non disponibile)
        self.history_path = Path('data/history')
        self.history_path.mkdir(parents=True, exist_ok=True)

        # Carica configurazione
        self.config = self._load_config()
    
    def _load_config(self) -> dict:
        """Carica configurazione dal file Excel"""
        try:
            df_config = pd.read_excel(self.excel_path, sheet_name='CONFIG', header=None)
            config = {}
            for _, row in df_config.iterrows():
                if pd.notna(row[0]) and pd.notna(row[1]):
                    key = str(row[0]).strip()
                    value = row[1]
                    config[key] = value
            return config
        except Exception as e:
            print(f"⚠️ Errore caricamento config: {e}")
            return {
                'Soglia RSI Ipervenduto': 30,
                'Soglia RSI Ipercomprato': 70,
                'Giorni Media Mobile': 15,
                'Min Indicatori Concordi L3': 2
            }
    
    def load_funds(self) -> pd.DataFrame:
        """
        Carica lista fondi dal file Excel

        Returns:
            DataFrame con tutti i fondi (ISIN validi, senza duplicati)
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Fondi')
            total_raw = len(df)

            # Rimuovi righe senza ISIN
            no_isin = df['ISIN'].isna()
            if no_isin.any():
                add_log(f"  ATTENZIONE: {no_isin.sum()} fondi senza ISIN ignorati: {df.loc[no_isin, 'Nome Fondo'].tolist()}")
                df = df[~no_isin]

            # Rimuovi ISIN duplicati (tieni prima occorrenza, logga i rimossi)
            dup_mask = df.duplicated('ISIN', keep='first')
            if dup_mask.any():
                dup_isins = df.loc[dup_mask, 'ISIN'].tolist()
                add_log(f"  ATTENZIONE: {dup_mask.sum()} righe duplicate rimosse (ISIN ripetuti): {dup_isins}")
                df = df[~dup_mask]

            add_log(f"📂 Caricati {len(df)} fondi validi (su {total_raw} righe nel file Excel)")
            return df.reset_index(drop=True)
        except Exception as e:
            print(f"❌ Errore caricamento fondi: {e}")
            return pd.DataFrame()
    
    def get_fund_history(self, isin: str) -> pd.Series:
        """
        Recupera storico prezzi per un fondo dal database PostgreSQL

        Args:
            isin: Codice ISIN del fondo

        Returns:
            Serie pandas con prezzi storici
        """
        # Recupera prezzo odierno (o la data fornita dalla fonte) e salvalo solo se è più recente
        nav_data = self.data_fetcher.get_nav(isin)
        if nav_data and nav_data.get('price') is not None:
            # Fonte può riportare la data del prezzo; altrimenti consideriamo oggi
            fetched_date_str = nav_data.get('date')
            try:
                fetched_date = datetime.fromisoformat(fetched_date_str).date() if fetched_date_str else datetime.now().date()
            except Exception:
                fetched_date = datetime.now().date()

            source = nav_data.get('source', 'FT Markets')

            # Controlla ultima data salvata nel DB (se disponibile)
            try:
                last_date_str = self.db.get_last_price_date(isin)
            except Exception:
                last_date_str = None

            last_date = None
            if last_date_str:
                try:
                    last_date = datetime.fromisoformat(last_date_str).date()
                except Exception:
                    last_date = None

            # Se la data recuperata non è più recente, saltiamo il salvataggio
            if last_date and fetched_date <= last_date:
                print(f"  ℹ️ Dato {fetched_date} non più recente di ultimo salvataggio {last_date}, skip save")
            else:
                # Salva nel database PostgreSQL; se fallisce, salva localmente in data/history
                try:
                    saved = self.db.save_price(isin, fetched_date.strftime('%Y-%m-%d'), nav_data['price'], source)
                except Exception:
                    saved = False

                if not saved:
                    # Fallback: append al file JSON locale per preservare lo storico
                    try:
                        history_file = self.history_path / f"{isin}.json"
                        history = []
                        if history_file.exists():
                            with open(history_file, 'r') as fh:
                                try:
                                    history = json.load(fh)
                                except Exception:
                                    history = []

                        history.append({'date': fetched_date.strftime('%Y-%m-%d'), 'price': nav_data['price'], 'source': source})
                        # Mantieni ultimi 365 giorni
                        if len(history) > 365:
                            history = history[-365:]
                        with open(history_file, 'w') as fh:
                            json.dump(history, fh)
                        print(f"  💾 Prezzo salvato in file locale: {isin} = {nav_data['price']} ({fetched_date.strftime('%Y-%m-%d')})")
                    except Exception as e:
                        print(f"❌ Errore salvataggio locale prezzo {isin}: {e}")

        # Recupera storico dal database (ultimi 100 giorni)
        prices = self.db.get_price_series(isin, days=100)

        if not prices.empty:
            return prices

        # Fallback: prova file JSON locale (per retrocompatibilità)
        history_file = self.history_path / f"{isin}.json"
        if history_file.exists():
            try:
                with open(history_file, 'r') as f:
                    history = json.load(f)
                if history:
                    df = pd.DataFrame(history)
                    return pd.Series(df['price'].values, index=pd.to_datetime(df['date']))
            except Exception:
                pass

        return pd.Series(dtype=float)
    
    def analyze_fund(self, row: pd.Series) -> dict:
        """
        Analizza un singolo fondo
        
        Args:
            row: Riga del DataFrame con dati fondo
        
        Returns:
            Dizionario con risultati analisi
        """
        isin = row['ISIN']
        level = int(row['Livello'])
        
        print(f"  📊 Analisi {row['Nome Fondo'][:40]}...")
        
        # Recupera storico dai file JSON locali (solo dati reali)
        prices = self.get_fund_history(isin)

        if len(prices) < 5:
            # Storico insufficiente - prova Yahoo Finance come ultima risorsa
            df_hist = self.data_fetcher.get_historical_nav(isin, days=45)
            if not df_hist.empty:
                prices = pd.Series(df_hist['nav'].values)
            # Se ancora insufficiente, l'analisi restituirà "dati insufficienti"
        
        # Rileva tipo di asset dalla categoria e crea analyzer dedicato
        categoria = row['Categoria']
        asset_type = TechnicalAnalyzer.detect_asset_type(categoria)
        analyzer = TechnicalAnalyzer(asset_type=asset_type)

        # Esegui analisi con parametri calibrati per il tipo di fondo
        analysis = analyzer.analyze_fund(prices, level=level)
        analysis['asset_type'] = asset_type

        # Analisi L0 "Deep Recovery" — valida sia per entry (level != 0) sia exit (level == 0)
        l0_analysis = analyzer.suggest_level_0(prices, level)

        return {
            'isin': isin,
            'nome': row['Nome Fondo'],
            'casa': row['Casa Gestione'],
            'categoria': categoria,
            'livello': level,
            'analysis': analysis,
            'l0_analysis': l0_analysis,
            'prices': prices,  # tenuto per calcolo panic_low in send_alerts
        }
    
    def update_excel(self, results: list):
        """
        Aggiorna il file Excel con i risultati dell'analisi
        Include aggiornamento automatico del livello

        Args:
            results: Lista di risultati analisi
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb['Fondi']

            # Mappa colonne Excel:
            # A(1)=Livello, B(2)=ISIN, C(3)=Nome, D(4)=Casa, E(5)=Categoria,
            # F(6)=Valuta, G(7)=Prezzo, H(8)=MM20, I(9)=RSI, J(10)=Segnale, K(11)=Ultima Modifica
            COL_LIVELLO = 1
            COL_ISIN = 2
            COL_PREZZO = 7
            COL_MM = 8
            COL_RSI = 9
            COL_SEGNALE = 10
            COL_ULTIMA_MODIFICA = 11

            # Mappa ISIN -> riga
            isin_to_row = {}
            for row in range(2, ws.max_row + 1):
                isin = ws.cell(row=row, column=COL_ISIN).value
                if isin:
                    isin_to_row[isin] = row

            level_changes = []

            # Aggiorna dati
            for result in results:
                isin = result['isin']
                analysis = result['analysis']

                if isin in isin_to_row:
                    row = isin_to_row[isin]

                    # Livello (colonna A) - AGGIORNAMENTO AUTOMATICO
                    current_level = result['livello']
                    suggested_level = analysis.get('suggested_level', current_level)
                    level_reason = analysis.get('level_reason', '')

                    if suggested_level != current_level:
                        ws.cell(row=row, column=COL_LIVELLO, value=suggested_level)
                        level_changes.append({
                            'nome': result['nome'],
                            'isin': isin,
                            'from': current_level,
                            'to': suggested_level,
                            'reason': level_reason
                        })
                        # Colora cella livello in base al cambio
                        level_cell = ws.cell(row=row, column=COL_LIVELLO)
                        if suggested_level < current_level:  # Upgrade (es. 3→2 o 2→1)
                            level_cell.fill = PatternFill("solid", fgColor="00B050")
                            level_cell.font = Font(bold=True, color="FFFFFF")
                        else:  # Downgrade (es. 1→2 o 2→3)
                            level_cell.fill = PatternFill("solid", fgColor="FF6600")
                            level_cell.font = Font(bold=True, color="FFFFFF")

                    # Prezzo (colonna G)
                    ws.cell(row=row, column=COL_PREZZO, value=analysis.get('current_price'))

                    # MM20 (colonna H)
                    ws.cell(row=row, column=COL_MM, value=analysis.get('ma'))

                    # RSI (colonna I)
                    ws.cell(row=row, column=COL_RSI, value=analysis.get('rsi'))

                    # Segnale (colonna J)
                    signal = analysis.get('final_signal', 'HOLD')
                    signal_cell = ws.cell(row=row, column=COL_SEGNALE, value=signal)

                    # Colora cella segnale
                    if signal == 'BUY':
                        signal_cell.fill = PatternFill("solid", fgColor="00B050")
                        signal_cell.font = Font(bold=True, color="FFFFFF")
                    elif signal == 'SELL':
                        signal_cell.fill = PatternFill("solid", fgColor="FF0000")
                        signal_cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        signal_cell.fill = PatternFill("solid", fgColor="FFC000")
                        signal_cell.font = Font(bold=True)

                    # Ultima Modifica (colonna K)
                    ws.cell(row=row, column=COL_ULTIMA_MODIFICA, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

            wb.save(self.excel_path)
            print(f"✅ File Excel aggiornato")

            # Log cambi di livello
            if level_changes:
                print(f"\n📊 CAMBI DI LIVELLO AUTOMATICI:")
                for change in level_changes:
                    arrow = "⬆️" if change['to'] < change['from'] else "⬇️"
                    print(f"  {arrow} {change['nome'][:40]}: L{change['from']} → L{change['to']}")
                    print(f"     Motivo: {change['reason']}")

            return level_changes

        except Exception as e:
            print(f"❌ Errore aggiornamento Excel: {e}")
            return []
    
    def generate_dashboard_data(self, results: list) -> dict:
        """
        Genera dati per la dashboard HTML
        
        Args:
            results: Lista risultati analisi
        
        Returns:
            Dizionario con dati dashboard
        """
        dashboard_data = {
            'last_update': datetime.now().isoformat(),
            'summary': {
                'total_funds': len(results),
                'buy_signals': 0,
                'sell_signals': 0,
                'hold_signals': 0
            },
            'levels': {
                1: [],
                2: [],
                3: []
            },
            'categories': {}
        }
        
        for r in results:
            signal = r['analysis'].get('final_signal', 'HOLD')
            # Usa il livello aggiornato (suggested_level) se è cambiato,
            # così i fondi promossi/declassati appaiono nel bucket corretto
            level = r['analysis'].get('suggested_level', r['livello'])
            category = r['categoria']
            
            # Conteggio segnali
            if signal == 'BUY':
                dashboard_data['summary']['buy_signals'] += 1
            elif signal == 'SELL':
                dashboard_data['summary']['sell_signals'] += 1
            else:
                dashboard_data['summary']['hold_signals'] += 1
            
            # Recupera prezzo di ieri dal database
            price_today = r['analysis'].get('current_price')
            price_yesterday = self.db.get_yesterday_price(r['isin'])

            # Calcola variazione percentuale (converti Decimal in float)
            change_pct = None
            if price_today and price_yesterday:
                change_pct = round(((float(price_today) - float(price_yesterday)) / float(price_yesterday)) * 100, 2)

            # Dati per livello
            fund_data = {
                'isin': r['isin'],
                'nome': r['nome'],
                'casa': r['casa'],
                'categoria': category,
                'price': float(price_today) if price_today else None,
                'price_yesterday': float(price_yesterday) if price_yesterday else None,
                'change_pct': change_pct,
                'ma': float(r['analysis'].get('ma')) if r['analysis'].get('ma') is not None else None,
                'rsi': float(r['analysis'].get('rsi')) if r['analysis'].get('rsi') is not None else None,
                'pct_1d': float(r['analysis'].get('pct_change_1d')) if r['analysis'].get('pct_change_1d') is not None else None,
                'pct_1w': float(r['analysis'].get('pct_change_1w')) if r['analysis'].get('pct_change_1w') is not None else None,
                'pct_1m': float(r['analysis'].get('pct_change_1m')) if r['analysis'].get('pct_change_1m') is not None else None,
                'buy_count': int(r['analysis'].get('buy_count', 0)),
                'asset_type': r['analysis'].get('asset_type', 'equity_developed'),
                'conditions': {
                    'allineamento_ok': bool(r['analysis'].get('level_conditions', {}).get('allineamento_ok', False)),
                    'persistenza_ok': bool(r['analysis'].get('level_conditions', {}).get('persistenza_ok', False)),
                    'rsi_optimal': bool(r['analysis'].get('level_conditions', {}).get('rsi_optimal', False)),
                    'distance_ok': bool(r['analysis'].get('level_conditions', {}).get('distance_ok', False)),
                    'adx_ok': bool(r['analysis'].get('level_conditions', {}).get('adx_ok', False)),
                    'adx_entry_ok': bool(r['analysis'].get('level_conditions', {}).get('adx_entry_ok', False)),
                    'nav_momentum_ok': bool(r['analysis'].get('level_conditions', {}).get('nav_momentum_ok', False)),
                },
                'signal_purity': r.get('signal_purity', {'available': False, 'reason': 'N/D'}),
            }
            dashboard_data['levels'][level].append(fund_data)

            # Dati per categoria
            if category not in dashboard_data['categories']:
                dashboard_data['categories'][category] = []
            dashboard_data['categories'][category].append(fund_data)
        
        # Salva JSON per dashboard (converte Decimal e numpy types)
        class SafeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                # Gestione tipi numpy
                import numpy as np
                if isinstance(obj, (np.integer,)):
                    return int(obj)
                if isinstance(obj, (np.floating,)):
                    return float(obj)
                if isinstance(obj, (np.bool_,)):
                    return bool(obj)
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                return super().default(obj)

        # Includi report di salute se disponibile
        if hasattr(self, '_health_report'):
            dashboard_data['health'] = self._health_report

        # Dati L0 "Deep Recovery" per la dashboard
        try:
            existing_l0 = self.db.get_all_l0_entries()
            today_d = datetime.now().date()
            l0_list = []
            for isin, entry in existing_l0.items():
                r = next((x for x in results if x['isin'] == isin), None)
                if not r:
                    continue
                price = r['analysis'].get('current_price')
                rsi = r['analysis'].get('rsi')
                l0a = r.get('l0_analysis', {})
                peak_val  = l0a.get('peak_price')
                peak_days = l0a.get('peak_days')
                dist_peak = l0a.get('distance_from_peak')
                pct_gain = None
                if price and entry['entry_price']:
                    pct_gain = round((float(price) - float(entry['entry_price'])) / float(entry['entry_price']) * 100, 2)
                try:
                    import numpy as np
                    ed = datetime.strptime(str(entry['entry_date'])[:10], '%Y-%m-%d').date()
                    days_in_l0 = max(1, int(np.busday_count(ed, today_d)) + 1)
                except Exception:
                    days_in_l0 = 1
                l0_list.append({
                    'isin': isin,
                    'nome': r['nome'],
                    'casa': r['casa'],
                    'categoria': r['categoria'],
                    'entry_date': str(entry['entry_date']),
                    'entry_price': float(entry['entry_price']),
                    'panic_low': float(entry['panic_low']),
                    'price': float(price) if price else None,
                    'rsi': float(rsi) if rsi else None,
                    'peak_price': float(peak_val) if peak_val else None,
                    'peak_days': peak_days,
                    'distance_from_peak': dist_peak,
                    'pct_gain': pct_gain,
                    'days_in_l0': days_in_l0,
                })
            dashboard_data['l0_funds'] = l0_list
        except Exception as e:
            dashboard_data['l0_funds'] = []
            print(f"Errore caricamento L0 per dashboard: {e}")

        with open('data/dashboard_data.json', 'w') as f:
            json.dump(dashboard_data, f, indent=2, cls=SafeEncoder)

        return dashboard_data
    
    def _compute_gap_analysis(self, result: dict) -> dict:
        """
        Calcola l'analisi quantitativa del gap tra un fondo L2 e i criteri L1 Pro.

        Per ciascuna delle 4 condizioni L1 Pro restituisce:
        - se è soddisfatta e il dettaglio
        - se non è soddisfatta: il gap numerico e la previsione

        Returns:
            Dizionario con buy_count, conditions (lista), e valori tecnici chiave
        """
        analysis = result['analysis']
        asset_type = analysis.get('asset_type', 'equity_developed')
        profile = TechnicalAnalyzer.PROFILES.get(asset_type, TechnicalAnalyzer.PROFILES['equity_developed'])

        lc = analysis.get('level_conditions', {})
        conditions = []

        # --- Condizione 1: TREND ---
        trend_ok = lc.get('trend_ok', False)
        days_above = lc.get('days_above_ma', 0)
        slope = float(lc.get('ma_slope', 0) or 0)
        distance = float(lc.get('distance_from_ma', 0) or 0)
        max_dist = profile['max_distance_from_ma']

        if trend_ok:
            conditions.append({
                'name': 'TREND', 'ok': True,
                'detail': f"Prezzo sopra MM da {days_above}gg, slope +{slope:.2f}%, distanza {distance:.1f}%",
                'gap_text': None, 'forecast': None
            })
        else:
            gaps, forecasts = [], []
            if not lc.get('price_above_ma_3days'):
                dn = max(0, 3 - days_above)
                gaps.append(f"Sopra MM da {days_above}/3 giorni richiesti")
                forecasts.append(f"Mancano {dn} seduta{'e' if dn != 1 else 'a'} consecutive sopra MM")
            if not lc.get('slope_positive'):
                gaps.append(f"Pendenza MM: {slope:.3f}% (serve > 0)")
                forecasts.append("1–2 giorni di prezzi crescenti possono invertire la pendenza MM")
            if not lc.get('distance_ok'):
                overshoot = distance - max_dist
                gaps.append(f"Distanza MM: {distance:.1f}% (max {max_dist:.1f}%, eccesso {overshoot:.1f}%)")
                forecasts.append(f"Il fondo ha corso troppo veloce: serve consolidamento di ~{overshoot:.1f}% o che la MM recuperi")
            conditions.append({
                'name': 'TREND', 'ok': False,
                'detail': None,
                'gap_text': ' · '.join(gaps) if gaps else 'Condizione non soddisfatta',
                'forecast': ' · '.join(forecasts) if forecasts else None
            })

        # --- Condizione 2: MOMENTUM (RSI) ---
        rsi_val = float(lc.get('rsi') or analysis.get('rsi') or 50)
        rsi_low = profile['rsi_optimal_low']
        rsi_high = profile['rsi_optimal_high']
        rsi_ok = lc.get('rsi_optimal', False)

        if rsi_ok:
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': True,
                'detail': f"RSI {rsi_val:.0f} nel range ottimale [{rsi_low}–{rsi_high}]",
                'gap_text': None, 'forecast': None
            })
        elif rsi_val < rsi_low:
            gap = rsi_low - rsi_val
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': False,
                'detail': None,
                'gap_text': f"RSI {rsi_val:.0f} sotto soglia {rsi_low} (mancano {gap:.0f} punti)",
                'forecast': f"Con {'2–3' if gap < 5 else '3–5'} sessioni positive, RSI può raggiungere la zona [{rsi_low}–{rsi_high}]"
            })
        else:
            gap = rsi_val - rsi_high
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': False,
                'detail': None,
                'gap_text': f"RSI {rsi_val:.0f} sopra soglia {rsi_high} (+{gap:.0f} punti, ipercomprato)",
                'forecast': "Probabile fase di consolidamento/correzione prima del rientro nel range ottimale"
            })

        # --- Condizione 3: VOLATILITÀ (Bande di Bollinger) ---
        bb_ok = lc.get('nav_above_upper_bb', False)
        price = float(analysis.get('current_price') or 0)
        bb = analysis.get('bollinger') or {}
        ma_val = float(analysis.get('ma') or 0)
        bb_upper = float(bb.get('upper') or 0)

        if bb_ok:
            conditions.append({
                'name': 'VOLATILITÀ (BB)', 'ok': True,
                'detail': "Prezzo nella metà superiore delle Bande di Bollinger",
                'gap_text': None, 'forecast': None
            })
        elif price and ma_val and bb_upper:
            if asset_type in TechnicalAnalyzer.EQUITY_FAMILY:
                midpoint = (ma_val + bb_upper) / 2
                gap_pct = (midpoint - price) / price * 100
                conditions.append({
                    'name': 'VOLATILITÀ (BB)', 'ok': False,
                    'detail': None,
                    'gap_text': f"Prezzo {price:.4f} sotto midpoint BB {midpoint:.4f} (gap: {gap_pct:.1f}%)",
                    'forecast': f"Serve +{gap_pct:.1f}% per entrare nella zona superiore BB: momentum ancora parzialmente sviluppato"
                })
            else:
                gap_pct = (ma_val - price) / price * 100 if price else 0
                conditions.append({
                    'name': 'VOLATILITÀ (BB)', 'ok': False,
                    'detail': None,
                    'gap_text': f"Prezzo {price:.4f} sotto MM {ma_val:.4f} (gap: {gap_pct:.1f}%)",
                    'forecast': f"Il NAV deve superare la MM ({ma_val:.4f}): manca un +{gap_pct:.1f}%"
                })
        else:
            conditions.append({
                'name': 'VOLATILITÀ (BB)', 'ok': False,
                'detail': None,
                'gap_text': "Dati Bollinger non disponibili (storico insufficiente)",
                'forecast': "Attendere accumulo di 20 giorni di storico per il calcolo delle BB"
            })

        # --- Condizione 4: SETUP (NAV in salita) — doppio criterio ---
        rising_days       = lc.get('rising_days', 0)
        nav_rising_orig   = lc.get('nav_rising_original', lc.get('nav_rising', False))
        nav_rising_alt    = lc.get('nav_rising_alt', False)
        pct_vs_5d         = lc.get('pct_vs_5d', 0.0)
        setup_ok          = lc.get('nav_rising', False)  # OR combinato

        sign_a  = '✅' if nav_rising_orig else '❌'
        sign_b  = '✅' if nav_rising_alt  else '❌'
        pct_str = f"{pct_vs_5d:+.2f}%" if pct_vs_5d != 0.0 else "N/D"
        # Formato HTML: ogni criterio su riga separata
        sub_a   = f"{sign_a} <b>A</b>: {rising_days} gg consecutivi in salita (soglia ≥3)"
        sub_b   = f"{sign_b} <b>B</b>: Prezzo vs 5gg fa {pct_str} (soglia >0%)"
        both    = f"{sub_a}<br>{sub_b}"

        if setup_ok:
            which = "A" if nav_rising_orig else "B"
            conditions.append({
                'name': 'SETUP (NAV)', 'ok': True,
                'detail': f"Criterio <b>{which}</b> passato<br>{sub_a}<br>{sub_b}",
                'gap_text': None, 'forecast': None
            })
        else:
            needed  = max(0, 3 - rising_days)
            fc_a    = f"<b>A</b>: mancano {needed} gg positivi consecutivi"
            fc_b    = (f"<b>B</b>: prezzo deve crescere {abs(pct_vs_5d):.2f}% rispetto a 5gg fa"
                       if pct_vs_5d <= 0 else f"<b>B</b>: prezzo già +{pct_vs_5d:.2f}% su 5gg (in recupero)")
            conditions.append({
                'name': 'SETUP (NAV)', 'ok': False,
                'detail': None,
                'gap_text': both,
                'forecast': f"{fc_a}. {fc_b}"
            })

        return {
            'buy_count': int(analysis.get('buy_count', 0)),
            'conditions': conditions,
            'price': price,
            'ma': ma_val,
            'rsi': rsi_val,
            'pct_1d': analysis.get('pct_change_1d'),
            'pct_1w': analysis.get('pct_change_1w'),
            'pct_1m': analysis.get('pct_change_1m'),
            'asset_type': asset_type,
        }

    def send_alerts(self, results: list):
        """
        Invia alert giornalieri basati sul tracking L1.

        - Email digest giornaliera con TUTTI i fondi in L1 (con prezzo entrata e % guadagno)
        - Email individuale per ogni fondo che ESCE da L1 (con prezzo entrata/uscita e %)
        """
        from datetime import date as date_type

        today = datetime.now().date()
        today_str = today.strftime('%Y-%m-%d')

        # Recupera tracking esistente dal DB
        existing_l1 = self.db.get_all_l1_entries()  # {isin: {entry_date, entry_price}}

        # Fondi con suggested_level == 1 oggi
        current_l1_isins = set()
        l1_funds_data = []

        for r in results:
            suggested = r['analysis'].get('suggested_level', r['livello'])
            if suggested != 1:
                continue

            isin = r['isin']
            price = r['analysis'].get('current_price')
            current_l1_isins.add(isin)

            # Nuovo ingresso in L1: registra data e prezzo
            if isin not in existing_l1:
                self.db.set_l1_entry(isin, today_str, price)
                entry_date = today
                entry_price = float(price) if price else None
                add_log(f"  🟢 Nuovo L1: {r['nome'][:40]} — entrato oggi a €{entry_price:.4f}" if entry_price else f"  🟢 Nuovo L1: {r['nome'][:40]}")
            else:
                entry = existing_l1[isin]
                entry_date = entry['entry_date']
                entry_price = entry['entry_price']

            # Calcola giorni in L1 (solo giorni di borsa, esclude weekend)
            try:
                import numpy as np
                ed = entry_date if isinstance(entry_date, date_type) else datetime.fromisoformat(str(entry_date)).date()
                days_in_l1 = max(1, int(np.busday_count(ed, today)) + 1)
            except Exception:
                days_in_l1 = 1

            pct_gain = None
            if price and entry_price:
                pct_gain = round((float(price) - float(entry_price)) / float(entry_price) * 100, 2)

            l1_funds_data.append({
                'isin': isin,
                'nome': r['nome'],
                'casa': r['casa'],
                'categoria': r['categoria'],
                'entry_date': entry_date,
                'entry_price': entry_price,
                'price': float(price) if price else None,
                'days_in_l1': days_in_l1,
                'pct_gain': pct_gain,
                'level_conditions': r['analysis'].get('level_conditions', {}),
            })

        # Fondi che escono da L1
        for isin, entry in existing_l1.items():
            if isin in current_l1_isins:
                continue

            # Trovalo nei risultati per il prezzo di uscita
            fund_result = next((r for r in results if r['isin'] == isin), None)
            if not fund_result:
                # Non è più nella lista fondi — rimuovi tracking silenziosamente
                self.db.remove_l1_entry(isin)
                continue

            exit_price = fund_result['analysis'].get('current_price')
            entry_price = entry['entry_price']
            entry_date = entry['entry_date']

            try:
                import numpy as np
                ed = entry_date if isinstance(entry_date, date_type) else datetime.fromisoformat(str(entry_date)).date()
                days_in_l1 = max(1, int(np.busday_count(ed, today)) + 1)
            except Exception:
                days_in_l1 = 1

            pct_gain = None
            if exit_price and entry_price:
                pct_gain = round((float(exit_price) - float(entry_price)) / float(entry_price) * 100, 2)

            sell_info = {
                'isin': isin,
                'nome': fund_result['nome'],
                'casa': fund_result['casa'],
                'categoria': fund_result['categoria'],
                'entry_date': entry_date,
                'entry_price': entry_price,
                'exit_price': float(exit_price) if exit_price else None,
                'days_in_l1': days_in_l1,
                'pct_gain': pct_gain,
                'conditions': fund_result['analysis'].get('level_conditions', {}),
            }
            exit_rule    = fund_result['analysis'].get('level_conditions', {}).get('exit_rule')
            exit_trigger = fund_result['analysis'].get('level_conditions', {}).get('exit_trigger', '')
            rule_label   = f'[Regola {exit_rule}: {exit_trigger}]' if exit_rule else ''
            pct_str      = f'{pct_gain:+.2f}%' if pct_gain is not None else 'N/D'
            add_log(f"  🔴 Uscita L1: {fund_result['nome'][:40]} — {pct_str} {rule_label}")
            self.alert_system.send_sell_l1_exit(sell_info)
            self.db.save_l1_exit(
                isin=isin,
                fund_name=fund_result['nome'],
                exit_date=today_str,
                exit_price=exit_price,
                exit_rule=exit_rule,
                exit_trigger=exit_trigger,
                entry_date=entry_date,
                entry_price=entry_price,
                days_in_l1=days_in_l1,
                pct_gain=pct_gain,
            )
            self.db.remove_l1_entry(isin)

        # Invia digest giornaliero L1
        if l1_funds_data:
            add_log(f"  📧 Invio digest L1: {len(l1_funds_data)} fondi in portafoglio")
            self.alert_system.send_l1_digest(l1_funds_data)
        else:
            add_log("  ℹ️ Nessun fondo in L1 — digest non inviato")

        # ── L0 "Deep Recovery" Tracking ──────────────────────────────────────
        existing_l0 = self.db.get_all_l0_entries()  # {isin: {entry_date, entry_price, panic_low}}

        current_l0_isins = set()
        l0_funds_data = []

        for r in results:
            isin = r['isin']
            l0 = r.get('l0_analysis', {})

            # Se il fondo è già in L0 tracking nel DB, gestisci uscite
            if isin in existing_l0:
                current_l0_isins.add(isin)
                entry = existing_l0[isin]
                entry_date = entry['entry_date']
                entry_price = entry['entry_price']
                panic_low = entry['panic_low']

                current_price = r['analysis'].get('current_price')
                current_rsi = r['analysis'].get('rsi')
                mm20 = r['analysis'].get('ma')

                # Calcola giorni in L0 (giorni di borsa)
                try:
                    import numpy as np
                    ed = entry_date if isinstance(entry_date, date_type) else datetime.fromisoformat(str(entry_date)).date()
                    days_in_l0 = max(1, int(np.busday_count(ed, today)) + 1)
                except Exception:
                    days_in_l0 = 1

                pct_gain = None
                if current_price and entry_price:
                    pct_gain = round((float(current_price) - float(entry_price)) / float(entry_price) * 100, 2)

                # Controlla regole di uscita
                exit_rule = None
                if current_price and panic_low and float(current_price) < float(panic_low):
                    exit_rule = 'alpha'
                elif current_rsi is not None and float(current_rsi) < 25:
                    exit_rule = 'beta'
                elif current_price and mm20 and float(current_price) > float(mm20):
                    exit_rule = 'gamma'
                elif days_in_l0 > 30:
                    exit_rule = 'epsilon'

                if exit_rule:
                    pct_str = f'{pct_gain:+.2f}%' if pct_gain is not None else 'N/D'
                    add_log(f"  🟠 Uscita L0 [{exit_rule}]: {r['nome'][:40]} — {pct_str}")
                    self.alert_system.send_sell_l0_exit({
                        'isin': isin,
                        'nome': r['nome'],
                        'casa': r['casa'],
                        'categoria': r['categoria'],
                        'entry_date': entry_date,
                        'entry_price': entry_price,
                        'panic_low': panic_low,
                        'exit_price': float(current_price) if current_price else None,
                        'days_in_l0': days_in_l0,
                        'pct_gain': pct_gain,
                        'exit_rule': exit_rule,
                        'conditions': {
                            **l0,
                            'mm20_current': mm20,
                            'rsi': current_rsi,
                        },
                    })
                    self.db.remove_l0_entry(isin)
                    current_l0_isins.discard(isin)
                else:
                    # Ancora in L0 — accumula per digest
                    l0_funds_data.append({
                        'isin': isin,
                        'nome': r['nome'],
                        'casa': r['casa'],
                        'categoria': r['categoria'],
                        'entry_date': entry_date,
                        'entry_price': entry_price,
                        'panic_low': panic_low,
                        'price': float(current_price) if current_price else None,
                        'days_in_l0': days_in_l0,
                        'pct_gain': pct_gain,
                        'level_conditions': {
                            **l0,
                            'mm20_current': mm20,
                            'rsi': current_rsi,
                        },
                    })
                continue

            # Fondo non ancora in L0: verifica se entra
            if not l0.get('l0_entry'):
                continue

            # Nuovo ingresso in L0
            current_price = r['analysis'].get('current_price')
            panic_low = l0.get('panic_low')

            if not current_price or not panic_low:
                continue

            self.db.set_l0_entry(isin, today_str, float(current_price), float(panic_low))
            current_l0_isins.add(isin)
            add_log(
                f"  🟠 Nuovo L0: {r['nome'][:40]} — "
                f"€{float(current_price):.4f} · panic_low=€{float(panic_low):.4f}"
            )
            l0_funds_data.append({
                'isin': isin,
                'nome': r['nome'],
                'casa': r['casa'],
                'categoria': r['categoria'],
                'entry_date': today,
                'entry_price': float(current_price),
                'panic_low': float(panic_low),
                'price': float(current_price),
                'days_in_l0': 1,
                'pct_gain': 0.0,
                'level_conditions': {
                    **l0,
                    'mm20_current': r['analysis'].get('ma'),
                    'rsi': r['analysis'].get('rsi'),
                },
            })

        # Fondi usciti dalla lista fondi (rimossi dall'Excel) ma ancora in L0 tracking
        for isin, entry in existing_l0.items():
            if isin in current_l0_isins:
                continue
            fund_result = next((r for r in results if r['isin'] == isin), None)
            if not fund_result:
                self.db.remove_l0_entry(isin)

        # Invia digest giornaliero L0
        if l0_funds_data:
            add_log(f"  📧 Invio digest L0: {len(l0_funds_data)} fondi in recupero")
            self.alert_system.send_l0_digest(l0_funds_data)
        else:
            add_log("  ℹ️ Nessun fondo in L0 — digest non inviato")

    def run(self, send_daily_report: bool = True):
        """
        Esegue ciclo completo di monitoraggio

        Args:
            send_daily_report: Se True, invia report giornaliero
        """
        import traceback
        add_log("="*50)
        add_log(f"FUND MONITOR - Avvio monitoraggio {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # 1. Carica fondi
        df_funds = self.load_funds()
        if df_funds.empty:
            add_log("ERRORE: Nessun fondo da monitorare (DataFrame vuoto)")
            return

        add_log(f"Caricati {len(df_funds)} fondi dal file Excel")

        # 1b. Fetch benchmark data per Signal Purity Score (una volta per run)
        benchmark_series = {}
        add_log("Fetching benchmark data (Signal Purity)...")
        try:
            fetched_tickers = {}
            for at, ticker in TechnicalAnalyzer.BENCHMARK_TICKERS.items():
                if ticker not in fetched_tickers:
                    s = self.data_fetcher.get_benchmark_history(ticker, days=120)
                    fetched_tickers[ticker] = s
                    add_log(f"  Benchmark {at} ({ticker}): {len(s)} gg" if not s.empty else f"  Benchmark {at} ({ticker}): N/D")
                benchmark_series[at] = fetched_tickers[TechnicalAnalyzer.BENCHMARK_TICKERS[at]]
        except Exception as e:
            add_log(f"  Benchmark fetch errore: {e}")

        # 2. Analizza ogni fondo
        add_log(f"Inizio analisi di {len(df_funds)} fondi...")
        results = []
        errors = []

        for idx, row in df_funds.iterrows():
            try:
                executor = ThreadPoolExecutor(max_workers=1)
                future = executor.submit(self.analyze_fund, row)
                timed_out = False
                try:
                    result = future.result(timeout=60)  # Max 60s per fondo
                except FuturesTimeoutError:
                    timed_out = True
                    add_log(f"  TIMEOUT {row['ISIN']} ({row['Nome Fondo'][:30]}): saltato dopo 60s")
                    errors.append({'isin': row['ISIN'], 'error': 'timeout 60s', 'traceback': ''})
                executor.shutdown(wait=False)  # Non bloccare — thread orfano continua in bg
                if timed_out:
                    continue
                # Signal Purity Score
                try:
                    asset_type = result['analysis'].get('asset_type', 'equity_developed')
                    bench = benchmark_series.get(asset_type)
                    if bench is not None and not bench.empty and not result['prices'].empty:
                        result['signal_purity'] = TechnicalAnalyzer.calculate_signal_purity(
                            result['prices'], bench
                        )
                    else:
                        result['signal_purity'] = {'available': False, 'reason': 'Benchmark N/D'}
                except Exception as pe:
                    result['signal_purity'] = {'available': False, 'reason': str(pe)}
                results.append(result)
                add_log(f"  OK {row['ISIN']} - {row['Nome Fondo'][:30]}")
                time.sleep(0.5)  # Rate limiting
            except Exception as e:
                error_detail = traceback.format_exc()
                errors.append({'isin': row['ISIN'], 'error': str(e), 'traceback': error_detail})
                add_log(f"  ERRORE {row['ISIN']}: {e}")
                add_log(f"  TRACEBACK: {error_detail}")

        add_log(f"Analisi completata: {len(results)} OK, {len(errors)} errori")

        # Salva report di salute per la dashboard e le email
        self._health_report = {
            'timestamp': datetime.now().isoformat(),
            'total_funds': len(df_funds),
            'funds_ok': len(results),
            'funds_error': len(errors),
            'errors': [{'isin': e['isin'], 'error': e['error']} for e in errors],
            'db_available': self.db.is_available() if self.db else False,
            'funds_with_price': sum(1 for r in results if r['analysis'].get('current_price') is not None),
            'funds_no_price': sum(1 for r in results if r['analysis'].get('current_price') is None),
        }

        # 3. Aggiorna Excel
        try:
            add_log("Step 3: Aggiornamento file Excel...")
            self.update_excel(results)
            add_log("Step 3: Excel aggiornato OK")
        except Exception as e:
            add_log(f"Step 3 ERRORE Excel: {e}")
            add_log(traceback.format_exc())

        # 4. Genera dati dashboard
        try:
            add_log(f"Step 4: Generazione dashboard con {len(results)} risultati...")
            os.makedirs('data', exist_ok=True)
            dashboard_data = self.generate_dashboard_data(results)
            total = dashboard_data.get('summary', {}).get('total_funds', '?')
            add_log(f"Step 4: Dashboard generata OK - {total} fondi")
        except Exception as e:
            add_log(f"Step 4 ERRORE Dashboard: {e}")
            add_log(traceback.format_exc())
            # Genera dashboard minima per non bloccare tutto
            dashboard_data = {
                'last_update': datetime.now().isoformat(),
                'summary': {'total_funds': len(results), 'buy_signals': 0, 'sell_signals': 0, 'hold_signals': 0},
                'levels': {1: [], 2: [], 3: []},
                'categories': {}
            }
            # Popola con dati minimi
            for r in results:
                try:
                    fund_data = {
                        'isin': r['isin'], 'nome': r['nome'], 'casa': r['casa'],
                        'categoria': r['categoria'], 'price': float(r['analysis'].get('current_price')) if r['analysis'].get('current_price') else None,
                        'price_yesterday': None, 'change_pct': None,
                        'ma': float(r['analysis'].get('ma')) if r['analysis'].get('ma') else None,
                        'rsi': float(r['analysis'].get('rsi')) if r['analysis'].get('rsi') else None,
                        'pct_1d': float(r['analysis'].get('pct_change_1d')) if r['analysis'].get('pct_change_1d') is not None else None,
                        'pct_1w': float(r['analysis'].get('pct_change_1w')) if r['analysis'].get('pct_change_1w') is not None else None,
                        'pct_1m': float(r['analysis'].get('pct_change_1m')) if r['analysis'].get('pct_change_1m') is not None else None,
                        'buy_count': int(r['analysis'].get('buy_count', 0)),
                        'asset_type': r['analysis'].get('asset_type', 'equity_developed'),
                        'conditions': {
                            'trend_ok': bool(r['analysis'].get('level_conditions', {}).get('trend_ok', False)),
                            'rsi_optimal': bool(r['analysis'].get('level_conditions', {}).get('rsi_optimal', False)),
                            'nav_above_bb': bool(r['analysis'].get('level_conditions', {}).get('nav_above_upper_bb', False)),
                            'nav_rising': bool(r['analysis'].get('level_conditions', {}).get('nav_rising', False)),
                            'nav_rising_original': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_original', False)),
                            'nav_rising_alt': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_alt', False)),
                        }
                    }
                    dashboard_data['levels'][r['livello']].append(fund_data)
                except:
                    pass
            with open('data/dashboard_data.json', 'w') as f:
                json.dump(dashboard_data, f, indent=2)
            add_log(f"Step 4: Dashboard fallback salvata con {len(results)} fondi")

        # 5. Invia alert
        try:
            add_log("Step 5: Invio alert...")
            self.send_alerts(results)
            add_log("Step 5: Alert OK")
        except Exception as e:
            add_log(f"Step 5 ERRORE Alert: {e}")

        # 6. Health report via email (solo in caso di errori)
        if hasattr(self, '_health_report'):
            try:
                add_log("Step 6: Invio health report...")
                self.alert_system.send_health_report(self._health_report)
                add_log("Step 6: Health report OK")
            except Exception as e:
                add_log(f"Step 6 ERRORE Health report: {e}")

        add_log(f"Monitoraggio completato - {datetime.now().strftime('%H:%M')}")
        add_log("="*50)


def main():
    """Entry point"""
    monitor = FundMonitor()
    monitor.run()


if __name__ == "__main__":
    main()
